import io
import html
import json
import math
import re
from datetime import datetime, date

import pandas as pd
import streamlit as st
from openai import OpenAI
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 頁面設定
# ============================================================

st.set_page_config(
    page_title="澳森托嬰中心｜觀察紀錄",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="collapsed",
)

DOMAINS = ["身體動作", "社會情緒", "語言溝通", "認知探索", "生活自理"]

OBS_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "results": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "record_id": {"type": "integer"},
                    "domain": {
                        "type": "string",
                        "enum": DOMAINS,
                    },
                    "observation": {"type": "string"},
                },
                "required": ["record_id", "domain", "observation"],
            },
        }
    },
    "required": ["results"],
}


# ============================================================
# 基礎工具
# ============================================================

def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def chinese_char_count(text):
    """
    計算「觀察紀錄」的大致中文字數。
    排除空白與換行，但保留中文、英文、數字與標點。
    """
    return len(re.sub(r"\s+", "", str(text or "")))


def normalize_name(name):
    return re.sub(r"\s+", "", clean_text(name))


def age_in_months(birthday, observation_date):
    """以觀察當日計算月齡。"""
    if pd.isna(birthday) or pd.isna(observation_date):
        return None

    try:
        b = pd.Timestamp(birthday)
        d = pd.Timestamp(observation_date)
    except Exception:
        return None

    months = (d.year - b.year) * 12 + (d.month - b.month)
    if d.day < b.day:
        months -= 1

    return max(months, 0)


def parse_age_text_to_months(age_text):
    """
    支援如 2Y2M19D、1Y11M16D。
    """
    s = clean_text(age_text).upper()
    if not s:
        return None

    m = re.search(r"(?:(\d+)Y)?(?:(\d+)M)?(?:(\d+)D)?", s)
    if not m:
        return None

    years = int(m.group(1) or 0)
    months = int(m.group(2) or 0)
    return years * 12 + months


# ============================================================
# Excel 讀取
# ============================================================

def find_sheet_with_columns(xls, required_columns):
    required = {c.strip() for c in required_columns}

    for sheet_name in xls.sheet_names:
        try:
            preview = pd.read_excel(xls, sheet_name=sheet_name, nrows=5)
            cols = {str(c).strip() for c in preview.columns}
            if required.issubset(cols):
                return sheet_name
        except Exception:
            continue

    return None


def load_uploaded_excel(uploaded_file):
    """
    必要資料：
    - 姓名
    - 日期
    - 老師的話

    若檔案中另有「姓名、生日、年齡」工作表，
    會自動計算每筆紀錄在當日的月齡，作為適齡判斷參考。
    """
    data = uploaded_file.getvalue()
    xls = pd.ExcelFile(io.BytesIO(data))

    source_sheet = find_sheet_with_columns(
        xls,
        ["姓名", "日期", "老師的話"],
    )

    if not source_sheet:
        raise ValueError(
            "找不到包含「姓名、日期、老師的話」三個欄位的工作表。"
        )

    source = pd.read_excel(xls, sheet_name=source_sheet)
    source.columns = [str(c).strip() for c in source.columns]

    source = source[["姓名", "日期", "老師的話"]].copy()
    source["姓名"] = source["姓名"].map(clean_text)
    source["老師的話"] = source["老師的話"].map(clean_text)
    source["日期"] = pd.to_datetime(source["日期"], errors="coerce")

    # 移除完全無法使用的列
    source = source[
        (source["姓名"] != "")
        & (source["老師的話"] != "")
        & source["日期"].notna()
    ].copy()

    source.reset_index(drop=True, inplace=True)
    source["record_id"] = range(1, len(source) + 1)

    # 嘗試找月齡資料
    age_sheet = None
    for sheet_name in xls.sheet_names:
        try:
            preview = pd.read_excel(xls, sheet_name=sheet_name, nrows=8)
            cols = {str(c).strip() for c in preview.columns}
            if "姓名" in cols and ("生日" in cols or "年齡" in cols):
                if sheet_name != source_sheet:
                    age_sheet = sheet_name
                    break
        except Exception:
            pass

    source["月齡"] = None

    if age_sheet:
        ages = pd.read_excel(xls, sheet_name=age_sheet)
        ages.columns = [str(c).strip() for c in ages.columns]

        keep = ["姓名"]
        if "生日" in ages.columns:
            keep.append("生日")
        if "年齡" in ages.columns:
            keep.append("年齡")

        ages = ages[keep].copy()
        ages["姓名"] = ages["姓名"].map(clean_text)
        ages = ages[ages["姓名"] != ""].copy()
        ages["_name_key"] = ages["姓名"].map(normalize_name)
        ages = ages.drop_duplicates("_name_key", keep="first")

        age_map = ages.set_index("_name_key").to_dict("index")

        def get_months(row):
            item = age_map.get(normalize_name(row["姓名"]))
            if not item:
                return None

            if "生日" in item and not pd.isna(item.get("生日")):
                months = age_in_months(item.get("生日"), row["日期"])
                if months is not None:
                    return months

            if "年齡" in item:
                return parse_age_text_to_months(item.get("年齡"))

            return None

        source["月齡"] = source.apply(get_months, axis=1)

    return source_sheet, age_sheet, source


# ============================================================
# OpenAI
# ============================================================

def get_client(api_key):
    key = (api_key or "").strip()
    if not key:
        raise ValueError("請先輸入 OpenAI API Key。")
    return OpenAI(api_key=key)


def make_ai_payload(batch_df):
    records = []

    for _, row in batch_df.iterrows():
        months = row.get("月齡")
        if pd.isna(months):
            months = None
        elif months is not None:
            months = int(months)

        records.append(
            {
                "record_id": int(row["record_id"]),
                "age_months": months,
                "teacher_note": row["老師的話"],
            }
        )

    return records


def generate_batch(client, model, batch_df):
    """
    不將幼兒姓名、生日傳給 AI。
    AI只收到 record_id、月齡及老師的話。
    """
    records = make_ai_payload(batch_df)

    instructions = """
你是臺灣托嬰中心的幼兒觀察紀錄整理助手。

請根據每一筆「老師的話」整理成一筆專業、客觀的幼兒觀察紀錄。

發展領域只能從以下五項選一項，選擇最主要、最能代表該筆觀察行為的領域：
1. 身體動作
2. 社會情緒
3. 語言溝通
4. 認知探索
5. 生活自理

撰寫規則：
- 使用繁體中文。
- 每筆觀察紀錄以約80～100個中文字為目標。
- 以「實際可觀察到的行為」為主，避免診斷、貼標籤、能力定論。
- 不可虛構原文沒有發生的行為、語句、情緒、成果或原因。
- 可將口語、表情符號、重複敘述整理成專業客觀文字。
- 優先聚焦課程活動、操作方式、互動、語言、探索、動作或自理行為。
- 午餐、午睡、喝奶等日常內容，只有在與本筆主要觀察領域有明確關聯時才納入。
- 不要寫「顯示其發展良好」、「能力優異」、「落後」等評量性語句。
- 若提供月齡，只作為活動與描述是否適齡的參考，不需要在觀察紀錄中寫出月齡。
- 每筆只能選一個發展領域。
- 請保持 record_id 與輸入一致。
"""

    response = client.responses.create(
        model=model,
        instructions=instructions,
        input=json.dumps(records, ensure_ascii=False),
        text={
            "format": {
                "type": "json_schema",
                "name": "daycare_observations",
                "schema": OBS_SCHEMA,
                "strict": True,
            }
        },
    )

    if not response.output_text:
        raise RuntimeError("AI 未回傳內容。")

    result = json.loads(response.output_text)
    return result["results"]


def revise_length(client, model, record_id, domain, observation, teacher_note):
    """
    若字數明顯偏短/偏長，單筆修正一次。
    """
    response = client.responses.create(
        model=model,
        instructions=(
            "你是托嬰中心觀察紀錄編輯。"
            "請在不增加原始紀錄未出現事實的前提下，"
            "將觀察紀錄調整為約80～100個中文字。"
            "保留原本發展領域，使用繁體中文、客觀可觀察語句。"
        ),
        input=json.dumps(
            {
                "record_id": int(record_id),
                "domain": domain,
                "current_observation": observation,
                "source_teacher_note": teacher_note,
            },
            ensure_ascii=False,
        ),
        text={
            "format": {
                "type": "json_schema",
                "name": "daycare_observation_revision",
                "schema": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "record_id": {"type": "integer"},
                        "domain": {"type": "string", "enum": DOMAINS},
                        "observation": {"type": "string"},
                    },
                    "required": ["record_id", "domain", "observation"],
                },
                "strict": True,
            }
        },
    )

    if not response.output_text:
        return domain, observation

    item = json.loads(response.output_text)
    return item["domain"], item["observation"]


def generate_all_observations(df, api_key, model, batch_size=8):
    client = get_client(api_key)

    results = []
    total = len(df)
    progress = st.progress(0, text="準備生成觀察紀錄…")
    status = st.empty()

    batches = math.ceil(total / batch_size)

    for batch_idx in range(batches):
        start = batch_idx * batch_size
        end = min(start + batch_size, total)
        batch_df = df.iloc[start:end]

        status.write(
            f"AI 生成中：第 {start + 1}～{end} 筆，共 {total} 筆"
        )

        batch_results = generate_batch(client, model, batch_df)
        result_map = {int(x["record_id"]): x for x in batch_results}

        for _, row in batch_df.iterrows():
            rid = int(row["record_id"])

            if rid not in result_map:
                raise RuntimeError(f"AI 遺漏第 {rid} 筆資料，請重新生成。")

            item = result_map[rid]
            domain = item["domain"]
            observation = item["observation"].strip()
            count = chinese_char_count(observation)

            # 約80-100字：若偏差過大，自動修一次
            if count < 75 or count > 110:
                domain, observation = revise_length(
                    client=client,
                    model=model,
                    record_id=rid,
                    domain=domain,
                    observation=observation,
                    teacher_note=row["老師的話"],
                )

            results.append(
                {
                    "record_id": rid,
                    "發展領域": domain,
                    "觀察紀錄": observation.strip(),
                }
            )

        progress.progress(
            end / total,
            text=f"已完成 {end} / {total} 筆",
        )

    status.empty()
    progress.empty()

    result_df = pd.DataFrame(results)

    merged = df.merge(result_df, on="record_id", how="left")
    merged["日期"] = pd.to_datetime(merged["日期"]).dt.date

    final_df = merged[
        ["姓名", "日期", "發展領域", "觀察紀錄"]
    ].copy()

    return final_df


# ============================================================
# Excel 匯出
# ============================================================

def export_excel_bytes(final_df):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        final_df.to_excel(
            writer,
            index=False,
            sheet_name="觀察紀錄",
        )

        ws = writer.book["觀察紀錄"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_gray = Side(style="thin", color="D9E2F3")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        widths = {
            "A": 14,
            "B": 14,
            "C": 14,
            "D": 65,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(
                    left=thin_gray,
                    right=thin_gray,
                    top=thin_gray,
                    bottom=thin_gray,
                )
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        # 日期欄
        for cell in ws["B"][1:]:
            cell.number_format = "yyyy/mm/dd"
            cell.alignment = Alignment(
                horizontal="center",
                vertical="top",
            )

        # 發展領域置中
        for cell in ws["C"][1:]:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
            )

        # 高度
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 52

        ws.sheet_view.showGridLines = False

    output.seek(0)
    return output.getvalue()


# ============================================================
# UI
# ============================================================

st.title("📝 澳森托嬰中心｜幼兒觀察紀錄")

st.markdown(
    "上傳老師紀錄 Excel 後，系統會讀取 **姓名、日期、老師的話**，"
    "若檔案中另有生日／年齡資料，也會自動計算月齡作為適齡參考。"
    "AI 最後只輸出 **發展領域＋80～100字左右觀察紀錄**。"
)

st.info(
    "發展領域固定為：身體動作、社會情緒、語言溝通、認知探索、生活自理。"
)

# ---------------------------
# 壹、上傳資料
# ---------------------------

st.subheader("壹、上傳 Excel")

uploaded_file = st.file_uploader(
    "請上傳 Excel 檔案",
    type=["xlsx", "xls"],
    help="至少需有：姓名、日期、老師的話。",
)

source_df = None

if uploaded_file is not None:
    try:
        source_sheet, age_sheet, source_df = load_uploaded_excel(uploaded_file)

        col1, col2, col3 = st.columns(3)
        col1.metric("讀取工作表", source_sheet)
        col2.metric("觀察資料筆數", len(source_df))
        col3.metric(
            "月齡資料",
            "已找到" if age_sheet else "未找到",
        )

        if age_sheet:
            st.caption(f"月齡參考工作表：{age_sheet}")

        preview = source_df[
            ["姓名", "日期", "老師的話", "月齡"]
        ].copy()
        preview["日期"] = preview["日期"].dt.strftime("%Y/%m/%d")
        preview["月齡"] = preview["月齡"].apply(
            lambda x: f"{int(x)}個月"
            if x is not None and not pd.isna(x)
            else ""
        )

        st.markdown("#### 原始資料預覽")
        st.dataframe(
            preview,
            use_container_width=True,
            hide_index=True,
            height=320,
        )

    except Exception as e:
        st.error(f"Excel 讀取失敗：{e}")
        source_df = None


# ---------------------------
# 貳、設定
# ---------------------------

if source_df is not None and len(source_df) > 0:
    st.divider()
    st.subheader("貳、生成設定")

    api_key = st.text_input(
        "OpenAI API Key",
        type="password",
        placeholder="請貼上你的 API Key",
        help=(
            "API Key 只用於本次 AI 呼叫。"
            "不寫入 app.py、Excel、GitHub、資料庫或網址參數。"
        ),
    )

    model = st.selectbox(
        "AI 模型",
        ["gpt-5-mini", "gpt-5.6", "gpt-4.1-mini"],
        index=0,
        help="一般大量觀察紀錄建議使用 gpt-5-mini，成本較低。",
    )

    st.caption(
        "🔒 為減少個資傳輸，AI 不會收到幼兒姓名或生日；"
        "只會收到流水編號、月齡與「老師的話」。"
    )

    if "observation_results" not in st.session_state:
        st.session_state.observation_results = None

    if st.button(
        f"✨ AI 生成 {len(source_df)} 筆觀察紀錄",
        type="primary",
        use_container_width=True,
    ):
        if not api_key.strip():
            st.error("請先輸入 OpenAI API Key。")
        else:
            try:
                with st.spinner("AI 正在整理觀察紀錄…"):
                    final_df = generate_all_observations(
                        source_df,
                        api_key,
                        model,
                    )

                st.session_state.observation_results = final_df
                st.success("觀察紀錄生成完成。")

            except Exception as e:
                st.error(f"生成失敗：{e}")


# ---------------------------
# 參、結果
# ---------------------------

if st.session_state.get("observation_results") is not None:
    final_df = st.session_state.observation_results

    st.divider()
    st.subheader("參、觀察紀錄結果")

    st.markdown("#### 表格預覽")

    display_df = final_df.copy()
    display_df["日期"] = pd.to_datetime(
        display_df["日期"]
    ).dt.strftime("%Y/%m/%d")

    # 直接組成「單行 HTML」交給 Streamlit，避免 Markdown
    # 把縮排後的 <tr>/<td> 誤判成程式碼區塊。
    css = """
    <style>
    .obs-table-wrap {
        width: 100%;
        overflow-x: hidden;
        margin-top: .4rem;
        margin-bottom: 1rem;
    }
    .obs-table {
        width: 100%;
        table-layout: fixed;
        border-collapse: collapse;
        font-size: 15px;
        line-height: 1.55;
    }
    .obs-table th {
        background: #1F4E78;
        color: #fff;
        font-weight: 700;
        text-align: center;
        padding: 9px 7px;
        border: 1px solid #D9E2F3;
    }
    .obs-table td {
        padding: 9px 7px;
        border: 1px solid #D9E2F3;
        vertical-align: top;
        white-space: normal;
        word-break: break-word;
        overflow-wrap: anywhere;
    }
    .obs-table th:nth-child(1), .obs-table td:nth-child(1) {
        width: 11%;
        text-align: center;
    }
    .obs-table th:nth-child(2), .obs-table td:nth-child(2) {
        width: 12%;
        text-align: center;
    }
    .obs-table th:nth-child(3), .obs-table td:nth-child(3) {
        width: 13%;
        text-align: center;
    }
    .obs-table th:nth-child(4), .obs-table td:nth-child(4) {
        width: 64%;
    }
    .obs-table tbody tr:nth-child(even) {
        background: #F7F9FC;
    }
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

    rows_html = []
    for _, row in display_df.iterrows():
        rows_html.append(
            "<tr>"
            f"<td>{html.escape(str(row['姓名']))}</td>"
            f"<td>{html.escape(str(row['日期']))}</td>"
            f"<td>{html.escape(str(row['發展領域']))}</td>"
            f"<td>{html.escape(str(row['觀察紀錄']))}</td>"
            "</tr>"
        )

    table_html = (
        '<div class="obs-table-wrap">'
        '<table class="obs-table">'
        '<thead><tr>'
        '<th>姓名</th>'
        '<th>日期</th>'
        '<th>發展領域</th>'
        '<th>觀察紀錄</th>'
        '</tr></thead>'
        '<tbody>'
        + "".join(rows_html)
        + '</tbody></table></div>'
    )

    st.markdown(table_html, unsafe_allow_html=True)

    st.divider()
    st.subheader("肆、下載 Excel")

    excel_bytes = export_excel_bytes(final_df)

    st.download_button(
        "📥 下載【幼兒觀察紀錄.xlsx】",
        data=excel_bytes,
        file_name="幼兒觀察紀錄.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
    )

    st.caption(
        "下載欄位固定為：姓名、日期、發展領域、觀察紀錄。"
    )
